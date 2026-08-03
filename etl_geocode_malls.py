#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Geocode 商场地址 → 经纬度（高德地图）
======================================
输入：商场数据结构化总表.xlsx → 商场总览 sheet
输出：output/malls_geocoded.json（地址→{lng,lat}缓存，增量更新）

用法：
  python etl_geocode_malls.py --key YOUR_AMAP_KEY --excel "商场数据结构化总表.xlsx" --output-dir ./output
  AMAP_KEY 也可通过环境变量 AMAP_KEY 传入
"""

import argparse
import json
import os
import sys
import time
import urllib.request
import urllib.parse

GEOCODE_URL = "https://restapi.amap.com/v3/geocode/geo"


def api_get(url):
    """调用高德API，带速率限制"""
    req = urllib.request.Request(url)
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read().decode("utf-8"))


def geocode_address(key, address, city=None):
    """正向地理编码：地址→经纬度"""
    params = {"key": key, "address": address}
    if city:
        params["city"] = city
    url = f"{GEOCODE_URL}?{urllib.parse.urlencode(params)}"
    try:
        data = api_get(url)
        if data.get("status") == "1" and data.get("geocodes"):
            gc = data["geocodes"][0]
            location = gc["location"]
            lng, lat = location.split(",")
            return float(lng), float(lat)
        else:
            print(f"[WARN] geocode failed: {address[:30]}... status={data.get('status')} {data.get('info')}")
            return None, None
    except Exception as e:
        print(f"[ERROR] geocode {address[:30]}... {e}")
        return None, None


def main():
    parser = argparse.ArgumentParser(description="Geocode mall addresses")
    parser.add_argument("--key", default=os.environ.get("AMAP_KEY", ""), help="高德 Web 服务 Key")
    parser.add_argument("--excel", required=True, help="商场数据结构化总表.xlsx 路径")
    parser.add_argument("--output-dir", default="./output")
    args = parser.parse_args()

    if not args.key:
        print("[ERROR] 缺少高德 Key，请用 --key 传入或设置环境变量 AMAP_KEY")
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)
    cache_file = os.path.join(args.output_dir, "malls_geocoded.json")

    # 加载缓存
    cache = {}
    if os.path.exists(cache_file):
        with open(cache_file, "r", encoding="utf-8") as f:
            cache = json.load(f)
        print(f"[INFO] 加载缓存: {len(cache)} 条已有坐标")

    # 读取 Excel
    import openpyxl
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    ws = wb["商场总览"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    new_count = 0
    skip_count = 0
    fail_count = 0
    total = ws.max_row - 1

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        city = str(row[0]).strip() if row[0] else ""
        name = str(row[1]).strip() if row[1] else ""
        address = str(row[2]).strip() if row[2] else ""

        cache_key = f"{city}_{name}"
        if cache_key in cache:
            cached = cache[cache_key]
            if cached.get("lng") and cached.get("lat"):
                skip_count += 1
                continue

        # 调高德 API (QPS 限 30/s，保守每秒 15 次)
        time.sleep(0.07)
        lng, lat = geocode_address(args.key, address, city)

        if lng is not None and lat is not None:
            cache[cache_key] = {"lng": lng, "lat": lat, "address": address}
            new_count += 1
            if new_count % 50 == 0:
                print(f"[PROGRESS] {new_count+skip_count}/{total} geocoded (new:{new_count} skip:{skip_count} fail:{fail_count})")
        else:
            fail_count += 1
            cache[cache_key] = {"lng": None, "lat": None, "address": address}

        # 每 200 条保存一次缓存
        if (new_count + fail_count) % 200 == 0:
            with open(cache_file, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)

    # 最终保存
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    print(f"\n[DONE] 总计: {total} 商场")
    print(f"  新增坐标: {new_count}")
    print(f"  缓存命中: {skip_count}")
    print(f"  失败: {fail_count}")
    print(f"  缓存文件: {cache_file}")


if __name__ == "__main__":
    main()
