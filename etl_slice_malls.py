#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 Excel + geocode 缓存 → 输出按城市+商场名称切片 JSON
=========================================================
输入：Excel 4个Sheet + malls_geocoded.json
输出：
  store-network-v2/public/data/malls/index.json        — 全量商场索引（名称/城市/经纬度/评分）
  store-network-v2/public/data/malls/{city}_{name}.json — 单个商场详情

用法：
  python etl_slice_malls.py --excel "商场数据结构化总表.xlsx" --geocode output/malls_geocoded.json --out-dir store-network-v2/public/data/malls
"""

import argparse
import json
import os
import re
import sys


def safe_num(v):
    """尝试转数字，失败返回 None"""
    if v is None or v == "" or v == "-":
        return None
    try:
        return float(str(v).replace(",", "").replace("万", "").replace("%", "").replace("㎡", "").strip())
    except:
        return None


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_demographic_pct(v):
    """解析百分比字符串 → 浮点数"""
    if v is None or v == "" or v == "-":
        return None
    s = str(v).replace("%", "").strip()
    try:
        return float(s)
    except:
        return None


def safe_filename(city, name):
    """安全文件名：移除特殊字符"""
    safe = re.sub(r'[\\/:*?"<>|#]', '_', f"{city}_{name}")
    return safe + ".json"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--geocode", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--max-restaurants", type=int, default=50, help="每家商场最多保存多少餐饮门店")
    args = parser.parse_args()

    # 加载 geocode 缓存
    with open(args.geocode, "r", encoding="utf-8") as f:
        geocode_cache = json.load(f)

    # 加载 Excel
    import openpyxl
    wb = openpyxl.load_workbook(args.excel, data_only=True)

    # ===== Sheet 1: 商场总览 → mall_map =====
    ws_main = wb["商场总览"]
    headers = [c.value for c in next(ws_main.iter_rows(min_row=1, max_row=1))]
    mall_map = {}  # key: (city, name) → dict

    for row in ws_main.iter_rows(min_row=2, values_only=True):
        city = safe_str(row[0])
        name = safe_str(row[1])
        address = safe_str(row[2])
        key = (city, name)

        if key in mall_map:
            continue  # 去重

        mall = {
            "name": name,
            "city": city,
            "address": address,
        }

        # Geo
        cache_key = f"{city}_{name}"
        gc = geocode_cache.get(cache_key, {})
        mall["lat"] = gc.get("lat")
        mall["lng"] = gc.get("lng")

        # Overview
        mall["overview"] = {
            "score": safe_num(row[3]),
            "score_rank": safe_str(row[4]),
            "population_score": safe_num(row[5]),
            "area_score": safe_num(row[6]),
            "consumption_score": safe_num(row[7]),
            "industry_score": safe_num(row[8]),
            "annual_sales": safe_str(row[9]),
            "annual_sales_unit": safe_str(row[10]),
            "area_size_sqm": safe_num(row[14]),  # 商业面积_万平米
            "open_date": safe_str(row[15]),
            "operator": safe_str(row[16]),
            "brand": safe_str(row[17]),
            "type": safe_str(row[18]),
            "floors": safe_str(row[19]) if not row[19] else f"{safe_str(row[19])}-{safe_str(row[20])}",
        }

        # Traffic
        mall["traffic"] = {
            "annual_daily": safe_num(row[11]),
            "annual_daily_unit": safe_str(row[12]),
            "annual_daily_rank": safe_str(row[13]),
            "jun_total": safe_num(row[21]),
            "jun_total_rank": safe_str(row[22]),
            "jun_daily_avg": safe_num(row[23]),
            "jun_daily_unit": safe_str(row[24]),
            "jun_daily_rank": safe_str(row[25]),
            "weekday_avg": safe_num(row[26]),
            "weekday_unit": safe_str(row[27]),
            "weekday_rank": safe_str(row[28]),
            "holiday_avg": safe_num(row[29]),
            "holiday_unit": safe_str(row[30]),
            "holiday_rank": safe_str(row[31]),
        }

        # Population (万人)
        mall["population"] = {
            "residential": {
                "500m": safe_num(row[32]),
                "1_5km": safe_num(row[33]),
                "3km": safe_num(row[34]),
            },
            "office": {
                "500m": safe_num(row[35]),
                "1_5km": safe_num(row[36]),
                "3km": safe_num(row[37]),
            },
            "permanent": {
                "500m": safe_num(row[38]),
                "1_5km": safe_num(row[39]),
                "3km": safe_num(row[40]),
            },
        }

        # Business types
        mall["business"] = {
            "competitors": safe_num(row[41]),
            "food": safe_num(row[42]),
            "shopping": safe_num(row[43]),
            "leisure": safe_num(row[44]),
            "education": safe_num(row[45]),
            "hotel": safe_num(row[46]),
            "services": safe_num(row[47]),
            "fitness": safe_num(row[48]),
            "auto": safe_num(row[49]),
            "total_stores": safe_num(row[59]),
            "old_store_3yr_pct": safe_num(row[60]),
        }

        # Nearby POI
        mall["nearby_poi"] = {
            "business": safe_num(row[50]),
            "community": safe_num(row[51]),
            "office": safe_num(row[52]),
            "school": safe_num(row[53]),
            "hospital": safe_num(row[54]),
            "transport": safe_num(row[55]),
            "scenic": safe_num(row[56]),
        }
        # C57 餐饮门店数, C58 周边商场数 (already in other sheets, skip)

        # Demographics
        mall["demographics"] = {
            "male_pct": parse_demographic_pct(row[61]),
            "female_pct": parse_demographic_pct(row[62]),
            "education": {
                "小学": parse_demographic_pct(row[63]),
                "初中": parse_demographic_pct(row[64]),
                "高中": parse_demographic_pct(row[65]),
                "大专": parse_demographic_pct(row[66]),
                "本科": parse_demographic_pct(row[67]),
                "硕士": parse_demographic_pct(row[68]),
                "博士": parse_demographic_pct(row[69]),
            },
            "has_children_pct": parse_demographic_pct(row[70]),
            "no_children_pct": parse_demographic_pct(row[71]),
        }

        # Scrape time
        mall["scraped_at"] = safe_str(row[72])

        mall_map[key] = mall

    print(f"[INFO] 商场总览: {len(mall_map)} 家")

    # ===== Sheet 2: 餐饮门店明细 → mall: nearby_restaurants =====
    ws_rest = wb["餐饮门店明细"]
    nearby_rest = {}  # key: (city, name) → list
    for row in ws_rest.iter_rows(min_row=2, values_only=True):
        city = safe_str(row[0])
        name = safe_str(row[1])
        key = (city, name)
        if key not in mall_map:
            continue
        if key not in nearby_rest:
            nearby_rest[key] = []
        if len(nearby_rest[key]) < args.max_restaurants:
            rest = {
                "brand": safe_str(row[2]),
                "address": safe_str(row[3]),
                "distance": safe_str(row[4]),
            }
            nearby_rest[key].append(rest)

    # Attach restaurants
    for key, rests in nearby_rest.items():
        mall_map[key]["nearby_restaurants"] = rests
    # Fill empty for malls with no restaurant data
    for key in mall_map:
        if "nearby_restaurants" not in mall_map[key]:
            mall_map[key]["nearby_restaurants"] = []

    print(f"[INFO] 餐饮门店: {sum(len(v) for v in nearby_rest.values())} 条匹配到 {len(nearby_rest)} 家商场")

    # ===== Sheet 3: 周边商场明细 → mall: nearby_malls =====
    ws_nm = wb["周边商场明细"]
    nearby_malls = {}  # key: (city, name) → list
    for row in ws_nm.iter_rows(min_row=2, values_only=True):
        city = safe_str(row[0])
        name = safe_str(row[1])
        key = (city, name)
        if key not in mall_map:
            continue
        if key not in nearby_malls:
            nearby_malls[key] = []
        nearby_malls[key].append({
            "name": safe_str(row[2]),
            "open_date": safe_str(row[3]),
            "area": safe_str(row[4]),
            "jun_daily_avg": safe_str(row[5]),
            "weekday_avg": safe_str(row[6]),
            "holiday_avg": safe_str(row[7]),
        })

    for key, ms in nearby_malls.items():
        mall_map[key]["nearby_malls"] = ms
    for key in mall_map:
        if "nearby_malls" not in mall_map[key]:
            mall_map[key]["nearby_malls"] = []

    print(f"[INFO] 周边商场: {sum(len(v) for v in nearby_malls.values())} 条匹配到 {len(nearby_malls)} 家商场")

    # ===== Sheet 4: 业态存续分析 → mall: business_survival =====
    ws_bs = wb["业态存续分析"]
    survival_map = {}
    for row in ws_bs.iter_rows(min_row=2, values_only=True):
        city = safe_str(row[0])
        name = safe_str(row[1])
        key = (city, name)
        if key not in mall_map:
            continue
        if key not in survival_map:
            survival_map[key] = []
        survival_map[key].append({
            "type": safe_str(row[2]),
            "count": safe_num(row[3]),
            "within_1yr_pct": parse_demographic_pct(row[4]),
            "1_3yr_pct": parse_demographic_pct(row[5]),
            "over_3yr_pct": parse_demographic_pct(row[6]),
        })

    for key, items in survival_map.items():
        mall_map[key]["business_survival"] = items
    for key in mall_map:
        if "business_survival" not in mall_map[key]:
            mall_map[key]["business_survival"] = []

    print(f"[INFO] 业态存续: {sum(len(v) for v in survival_map.values())} 条匹配到 {len(survival_map)} 家商场")

    # ===== 输出 =====
    os.makedirs(args.out_dir, exist_ok=True)

    # Index
    index = []
    write_count = 0
    for (city, name), mall in mall_map.items():
        index.append({
            "name": name,
            "city": city,
            "lat": mall["lat"],
            "lng": mall["lng"],
            "score": mall["overview"]["score"],
            "type": mall["overview"]["type"],
            "open_date": mall["overview"]["open_date"],
        })

        # Per-mall JSON
        fname = safe_filename(city, name)
        fpath = os.path.join(args.out_dir, fname)
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(mall, f, ensure_ascii=False, indent=2)
        write_count += 1

    # Write index
    index_path = os.path.join(args.out_dir, "index.json")
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False)

    # Write manifest for partial loading
    manifest = {}
    for m in index:
        city = m["city"]
        if city not in manifest:
            manifest[city] = []
        manifest[city].append({
            "name": m["name"],
            "file": safe_filename(city, m["name"]),
            "lat": m["lat"],
            "lng": m["lng"],
            "score": m["score"],
        })
    manifest_path = os.path.join(args.out_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False)

    print(f"\n[DONE]")
    print(f"  Index: {index_path} ({len(index)} 条)")
    print(f"  Manifest: {manifest_path} ({len(manifest)} 城市)")
    print(f"  Per-mall JSON: {write_count} 个文件 → {args.out_dir}/")

    # Size stats
    total_size = 0
    for f in os.listdir(args.out_dir):
        if f.endswith(".json"):
            total_size += os.path.getsize(os.path.join(args.out_dir, f))
    print(f"  Total size: {total_size/1024/1024:.1f} MB")


if __name__ == "__main__":
    main()
